import os
import sys
sys.path.append('../')
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from utils_ken.log.getlog import get_logger
from utilslegend.utils_mode_counter import POLYGON_COUNTER_IDS
from utilslegend.utils_mode_counter import COUNTER_TWOWAY_IDS

logexcel         = get_logger('excel','./logs/excel.log')


def get_datetime() :
    currentDT = datetime.datetime.now().strftime("%Y/%m/%d")
    return currentDT

def get_hour() :
    currentHour = datetime.datetime.now().hour
    return currentHour

class export_excel() :
    def __init__(self,excel_fille, camera_id, camera_name) :
        self.excel_file     = excel_fille
        self.camera_id      = camera_id
        self.camera_name    = camera_name

        self.sheet          = {}

        if os.path.isfile(self.excel_file) :
            self.load_file()
        else :
            self.create_file()
    
    def load_file(self) :
        logexcel.info("load_file : {}".format(self.excel_file))
        try :
            self.book   = load_workbook(self.excel_file)
            sheetnames  =  self.book.sheetnames
            for sheet_name in sheetnames :

                self.sheet[sheet_name] = self.book.get_sheet_by_name(sheet_name)
            # self.sheet = self.book.active
        except :
            self.create_file()

    def create_sheet(self, sheet_name, id_func, func_name) :
        # print(self.book.sheetnames)
        sheetnames = self.book.sheetnames
        if sheet_name not in sheetnames :
            self.sheet[sheet_name] = self.book.create_sheet(sheet_name,0)

            self.sheet[sheet_name]['A1'] = 'Cam Name : '
            self.sheet[sheet_name]['B1'] = self.camera_name

            self.sheet[sheet_name]['A2'] = 'ID :'
            self.sheet[sheet_name]['B2'] = id_func
            
            self.sheet[sheet_name]['A3'] = 'Line Name :'
            self.sheet[sheet_name]['B3'] = func_name

            self.sheet[sheet_name]['A5'] = 'Time'
            
            if id_func in COUNTER_TWOWAY_IDS :
                self.sheet[sheet_name]['B5'] = 'Car'
                self.sheet[sheet_name]['D5'] = 'Truck'
                self.sheet[sheet_name]['F5'] = 'Bus'
                self.sheet[sheet_name]['H5'] = 'Bike'
                self.sheet[sheet_name]['J5'] = 'Person'

                self.sheet[sheet_name]['B6'] = 'up'
                self.sheet[sheet_name]['C6'] = 'down'

                self.sheet[sheet_name]['D6'] = 'up'
                self.sheet[sheet_name]['E6'] = 'down'

                self.sheet[sheet_name]['F6'] = 'up'
                self.sheet[sheet_name]['G6'] = 'down'

                self.sheet[sheet_name]['H6'] = 'up'
                self.sheet[sheet_name]['I6'] = 'down'
                
                self.sheet[sheet_name]['J6'] = 'up'
                self.sheet[sheet_name]['K6'] = 'down'

            if id_func in POLYGON_COUNTER_IDS :

                self.sheet[sheet_name]['B5'] = 'Car'
                self.sheet[sheet_name]['C5'] = 'Truck'
                self.sheet[sheet_name]['D5'] = 'Bus'
                self.sheet[sheet_name]['E5'] = 'Bike'
                self.sheet[sheet_name]['F5'] = 'Person'

        else :
            logexcel.info("shet : {} exits".format(sheet_name))
            
            # pass 

    def create_file(self) :
        logexcel.info("create_file : {}".format(self.excel_file))
        self.book   = Workbook()
        # self.sheet  = self.book.active
        # self.sheet['A1'] = 'Date/Hour'
        # row     = 1
        # column  = 4
        # for i in range(24) :
        #     self.sheet.cell(row=row, column=column).value = i
        #     column +=1

        # self.current_row = 2

    def update_line(self, sheet_name, list_counter_up, list_counter_dow) :

        # print(self.sheet)
        current_row     = self.get_current_row(sheet_name)
        current_column  = 2
        # print(current_row)
        hour    = get_hour()
        label_time = '{}:00'.format(hour)
        # print(label_time)
        value_check  = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value
        if value_check is None :
            self.sheet[sheet_name].cell(row=current_row, column=current_column - 1 ).value       = label_time

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = list_counter_up[0]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value       = list_counter_dow[0]

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 2).value       = list_counter_up[1]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 3).value       = list_counter_dow[1]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 4).value       = list_counter_up[2]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 5).value       = list_counter_dow[2]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 6).value       = list_counter_up[3]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 7).value       = list_counter_dow[3]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 8).value       = list_counter_up[4]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 9).value       = list_counter_dow[4]

        else :

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value + list_counter_up[0]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value + list_counter_dow[0]

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 2).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 2).value + list_counter_up[1]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 3).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 3).value + list_counter_dow[1]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 4).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 4).value + list_counter_up[2]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 5).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 5).value + list_counter_dow[2]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 6).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 6).value + list_counter_up[3]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 7).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 7).value + list_counter_dow[3]


            self.sheet[sheet_name].cell(row=current_row, column=current_column + 8).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 8).value + list_counter_up[4]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 9).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 9).value + list_counter_dow[4]


    def update_polygon(self, sheet_name, list_counter_up) :
        current_row     = self.get_current_row(sheet_name)
        current_column  = 2
        # print(current_row)
        hour    = get_hour()
        label_time = '{}:00'.format(hour)
        # print(label_time)

        self.sheet[sheet_name].cell(row=current_row, column=current_column - 1 ).value       = label_time

        self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = list_counter_up[0]
        self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value       = list_counter_up[1]
        self.sheet[sheet_name].cell(row=current_row, column=current_column + 2).value       = list_counter_up[2]
        self.sheet[sheet_name].cell(row=current_row, column=current_column + 3).value       = list_counter_up[3]
        self.sheet[sheet_name].cell(row=current_row, column=current_column + 4).value       = list_counter_up[4]



    def get_current_row(self, sheet_name) :
        current_time = get_datetime()
        column  = 2
        row     = 6
        while True :
            cellCR = self.sheet[sheet_name].cell(row=row, column=column).value
            if cellCR == None:
                current_row = row
                break
            
            else : 
                print(cellCR)
                print(type(cellCR))

            row += 1

        return current_row

    # def get_current_column(self) :
    #     current_hour = get_hour()
    #     current_column = int(current_hour) + 4
    #     return current_column


    def save(self) :
        self.book.save(self.excel_file)


def run() :
    now = datetime.datetime.now()
    date_folder = now.strftime('%Y-%m-%d')
    # logexcel.info(date_folder)
    excel_file                  = "cam_{}_{}_{}.xlsx".format(1,"helo",date_folder)
    camera_id   = 10
    
    camera_Name   = 'nguyentho'

    sheet_name      = 'line_id1'
    sheet_name2     = 'line_id2'
    id_func         = 5
    id_func2        = 12
    func_name       = 'abc'
    excel = export_excel(excel_file, camera_id, camera_Name)

    # print(excel.sheetnames)
    excel.create_sheet(sheet_name, id_func, func_name)
    excel.create_sheet(sheet_name2, id_func2, func_name)

    list_counter_up     = [1,2,3,4,5]
    list_counter_dow    = [6,7,8,9,1]
    # time_update         = 45
    excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 60
    excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 65
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 70
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 830
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    excel.update_polygon(sheet_name2,  list_counter_up)
    excel.update_polygon(sheet_name2,  list_counter_up)
    # excel.update_polygon(sheet_name2,  list_counter_up)
    # excel.update_polygon(sheet_name2,  list_counter_up)
    # # excel.update(10)

    print(excel.sheet)
    excel.save()


if __name__ == '__main__' :
    run()