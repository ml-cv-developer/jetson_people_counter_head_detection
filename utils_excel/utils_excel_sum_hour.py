import os
import sys
sys.path.append('../')
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from utils_ken.log.getlog import get_logger
from utils.utils_mode_counter import POLYGON_COUNTER_IDS
from utils.utils_mode_counter import COUNTER_ONEWAY_IDS

logexcel         = get_logger('excel','./logs/excel.log')


def get_datetime() :
    currentDT = datetime.datetime.now().strftime("%Y/%m/%d")
    return currentDT

# def rdf_get_datetime() :
#     now = datetime.datetime.now()
#     # str_time = now.strftime('%Y-%m-%d %H:%M:%S')
#     str_time = (datetime.datetime.now() - datetime.timedelta(minutes=2)).strftime('%Y-%m-%d %H:%M:%S')
#     # legend_db.info(str_time)
#     return str_time


def get_hour() :
    currentHour = (datetime.datetime.now() - datetime.timedelta(minutes=2)).hour
    return currentHour

def hour_from_test(hour_test) :
    hout_int = int(hour_test.split(":")[0])
    return hout_int

class export_excel() :
    def __init__(self,excel_fille, camera_id, camera_name) :
        self.excel_file     = excel_fille
        self.camera_id      = camera_id
        self.camera_name    = camera_name

        self.sheet          = {}

        if os.path.isfile(self.excel_file) :
            self.load_file()
        else :
            self.create_file()
    
    def load_file(self) :
        logexcel.info("load_file : {}".format(self.excel_file))
        try :
            self.book   = load_workbook(self.excel_file)
            sheetnames  =  self.book.sheetnames
            for sheet_name in sheetnames :
                self.sheet[sheet_name] = self.book.get_sheet_by_name(sheet_name)
            # self.sheet = self.book.active
        except :
            self.create_file()

    def create_sheet(self, sheet_name, id_func, func_name) :
        # print(self.book.sheetnames)
        logexcel.info("create_sheet : {}".format(sheet_name))
        sheetnames = self.book.sheetnames
        if sheet_name not in sheetnames :
            self.sheet[sheet_name] = self.book.create_sheet(sheet_name,0)

            self.sheet[sheet_name]['A1'] = 'Cam Name : '
            self.sheet[sheet_name]['B1'] = self.camera_name

            self.sheet[sheet_name]['A2'] = 'ID :'
            self.sheet[sheet_name]['B2'] = id_func
            
            self.sheet[sheet_name]['A3'] = 'Line Name :'
            self.sheet[sheet_name]['B3'] = func_name

            self.sheet[sheet_name]['A5'] = 'Time'
            
            if id_func in COUNTER_ONEWAY_IDS :
                self.sheet[sheet_name]['B5'] = 'Person'

                self.sheet[sheet_name]['B6'] = 'up'
                self.sheet[sheet_name]['C6'] = 'down'


            if id_func in POLYGON_COUNTER_IDS :
                self.sheet[sheet_name]['B5'] = 'Person'

        else :
            logexcel.info("sheet : {} exits".format(sheet_name))
            
            # pass 

    def create_file(self) :
        logexcel.info("create_file : {}".format(self.excel_file))
        self.book   = Workbook()
        # self.current_row = 2

    def update_line(self, sheet_name, list_counter_up, list_counter_dow) :
        logexcel.info("Update Up {} : {}".format(sheet_name, list_counter_up))
        logexcel.info("Update Down {} : {}".format(sheet_name, list_counter_dow))
        # print(self.sheet)
        current_row     = self.get_current_row(sheet_name)
        current_column  = 2
        # print(current_row)
        hour    = get_hour()
        label_time = '{}:00'.format(hour)
        # print(label_time)
        value_check  = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value
        if value_check is None :
            self.sheet[sheet_name].cell(row=current_row, column=current_column - 1 ).value       = label_time

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = list_counter_up[0]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value       = list_counter_dow[0]

        else :

            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value + list_counter_up[0]
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 1).value + list_counter_dow[0]


    def update_polygon(self, sheet_name, list_counter_up) :
        logexcel.info("Update po {} : {}".format(sheet_name, list_counter_up))
        current_row     = self.get_current_row(sheet_name)
        current_column  = 2
        # print(current_row)
        hour    = get_hour()
        label_time = '{}:00'.format(hour)
        # print(label_time)

        self.sheet[sheet_name].cell(row=current_row, column=current_column - 1 ).value       = label_time

        value_check  = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value
        if value_check is None :
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = list_counter_up[0]

        else :
            self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value       = self.sheet[sheet_name].cell(row=current_row, column=current_column + 0).value + list_counter_up[0]




    def get_current_row(self, sheet_name) :
        current_hour    = get_hour()
        column  = 2
        row     = 6
        while True :
            cellCR      = self.sheet[sheet_name].cell(row=row, column=column).value
            cell_HOUR   = self.sheet[sheet_name].cell(row=row, column=1).value
            if cellCR is None:
                current_row = row
                break
            
            if cell_HOUR is not None :
                hour_int = hour_from_test(cell_HOUR)
                if current_hour == hour_int :
                    current_row = row
                    # print("current_row :{}".format(current_row))
                    break

            row += 1

        return current_row

    # def get_current_column(self) :
    #     current_hour = get_hour()
    #     current_column = int(current_hour) + 4
    #     return current_column


    def save(self) :
        self.book.save(self.excel_file)


def run() :
    now = datetime.datetime.now()
    date_folder = now.strftime('%Y-%m-%d')
    # logexcel.info(date_folder)
    excel_file                  = "cam_{}_{}_{}.xlsx".format(1,"helo",date_folder)
    camera_id   = 10
    
    camera_Name   = 'nguyentho'

    sheet_name      = 'line_id1'
    sheet_name2     = 'line_id2'
    id_func         = 5
    id_func2        = 12
    func_name       = 'abc'
    excel = export_excel(excel_file, camera_id, camera_Name)

    # print(excel.sheetnames)
    excel.create_sheet(sheet_name, id_func, func_name)
    excel.create_sheet(sheet_name2, id_func2, func_name)

    list_counter_up     = [1,2,3,4,5]
    list_counter_dow    = [6,7,8,9,1]
    # time_update         = 45
    excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 60
    excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 65
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 70
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    # # time_update         = 830
    # excel.update_line(sheet_name,  list_counter_up, list_counter_dow)
    excel.update_polygon(sheet_name2,  list_counter_up)
    excel.update_polygon(sheet_name2,  list_counter_up)
    # excel.update_polygon(sheet_name2,  list_counter_up)
    # excel.update_polygon(sheet_name2,  list_counter_up)
    # # excel.update(10)

    # print(excel.sheet)
    excel.save()


if __name__ == '__main__' :
    run()